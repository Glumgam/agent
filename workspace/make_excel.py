import openpyxl
from random import randint
date = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
sheet = openpyxl.Workbook()
sheet.active.title = 'Sales'
for i in range(6):
    sheet.cell(row=i+1, column=1).value = date[i]
    sheet.cell(row=i+1, column=2).value = randint(10000, 50000)
    sheet.cell(row=i+1, column=3).value = '=B' + str(i+1) + '/B2'
sheet.save('sales.xlsx')