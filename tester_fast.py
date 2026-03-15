#!/usr/bin/env python3
"""
tester_fast.py — 軽量テストランナー

- Pythonで直接実行できるテストはエージェントを使わない
- エージェントが必要なテスト(Coding)のみLLMをサブプロセス経由で呼ぶ
- 1テストあたりのタイムアウト: 60秒
- 失敗のみ最大1回リトライ
"""

import json
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

WORKSPACE = Path("/Volumes/ESD-EHA/agent/workspace")
AGENT_ROOT = Path("/Volumes/ESD-EHA/agent")
VENV_PYTHON = str(AGENT_ROOT / "venv" / "bin" / "python")
MAIN_PY = str(AGENT_ROOT / "main.py")
TIMEOUT = 300  # 秒/テスト (Ollama モデルロード込みで最大5分)


# =====================================================
# カテゴリ1: コーディング（エージェント使用、max_steps=5）
# =====================================================

CODING_TESTS = [
    {
        "id": "C1",
        "label": "FizzBuzz",
        "task": (
            "fizzbuzz.pyを作成して1から15までのFizzBuzzを出力せよ。"
            "実行して出力を確認したら即座に done を宣言せよ。"
        ),
        "expect": "FizzBuzz",
        "max_steps": 5,
        "use_agent": True,
    },
    {
        "id": "C2",
        "label": "バグ修正",
        "task": (
            "bugfix.pyにdef add(a,b): return a-b を書き、"
            "a+bに修正してadd(3,4)を実行し7を出力せよ。"
            "出力を確認したら即座に done を宣言せよ。"
        ),
        "expect": "7",
        "max_steps": 5,
        "use_agent": True,
    },
]


# =====================================================
# カテゴリ2: ファイル操作（直接Python実行）
# =====================================================

def test_F1():
    """拡張子別フォルダ振り分け"""
    import shutil
    tmp = WORKSPACE / "sort_test"
    if tmp.exists():
        shutil.rmtree(tmp)
    tmp.mkdir(parents=True)
    (tmp / "a.txt").write_text("hello")
    (tmp / "b.py").write_text("print(1)")
    (tmp / "c.md").write_text("# doc")
    for ext, folder in [(".txt", "texts"), (".py", "scripts"), (".md", "docs")]:
        d = tmp / folder
        d.mkdir(exist_ok=True)
        for f in list(tmp.glob(f"*{ext}")):
            if f.name.startswith("."):  # macOS ._xxx 隠しファイルをスキップ
                continue
            shutil.move(str(f), str(d / f.name))
    assert (tmp / "texts" / "a.txt").exists()
    assert (tmp / "scripts" / "b.py").exists()
    assert (tmp / "docs" / "c.md").exists()
    shutil.rmtree(tmp)
    return "OK: ファイル振り分け成功"


def test_F2():
    """重複ファイル検出・削除"""
    import hashlib, shutil
    tmp = WORKSPACE / "dup_test"
    if tmp.exists():
        shutil.rmtree(tmp)
    tmp.mkdir(parents=True)
    for name in ["dup1.txt", "dup2.txt", "dup3.txt"]:
        (tmp / name).write_text("same content")
    (tmp / "unique.txt").write_text("different content")
    hashes: dict = {}
    deleted = []
    for f in sorted(tmp.iterdir()):
        if f.name.startswith("."):  # macOS ._xxx 隠しファイルをスキップ
            continue
        h = hashlib.md5(f.read_bytes()).hexdigest()
        if h in hashes:
            f.unlink()
            deleted.append(f.name)
        else:
            hashes[h] = f
    assert len(deleted) == 2, f"expected 2 deleted, got {len(deleted)}"
    shutil.rmtree(tmp)
    return f"OK: 重複 {len(deleted)} 件を削除"


def test_F3():
    """Pythonファイル行数集計"""
    files = sorted(AGENT_ROOT.glob("*.py"))[:5]
    total = sum(len(f.read_text(errors="ignore").splitlines()) for f in files)
    assert total > 0
    return f"OK: {len(files)} ファイル, 合計 {total} 行"


def test_F4():
    """ログファイル解析"""
    import re
    log_path = WORKSPACE / "sample.log"
    WORKSPACE.mkdir(exist_ok=True)
    lines = [
        "[ERROR] connection refused",
        "[WARNING] slow response",
        "[INFO] startup complete",
        "[ERROR] disk full",
        "[INFO] job done",
    ]
    log_path.write_text("\n".join(lines))
    counts: dict = {}
    for ln in log_path.read_text().splitlines():
        m = re.search(r"\[(ERROR|WARNING|INFO)\]", ln)
        if m:
            counts[m.group(1)] = counts.get(m.group(1), 0) + 1
    assert "ERROR" in counts and counts["ERROR"] == 2
    return f"OK: ERROR={counts.get('ERROR',0)} WARNING={counts.get('WARNING',0)} INFO={counts.get('INFO',0)}"


FILE_TESTS = [
    {"id": "F1", "label": "拡張子別振り分け", "fn": test_F1},
    {"id": "F2", "label": "重複ファイル検出", "fn": test_F2},
    {"id": "F3", "label": "行数集計",          "fn": test_F3},
    {"id": "F4", "label": "ログ解析",           "fn": test_F4},
]


# =====================================================
# カテゴリ3: PDF（直接Python実行）
# =====================================================

def test_P1():
    """PDF生成"""
    try:
        from reportlab.pdfgen import canvas as rl_canvas
    except ImportError:
        return "SKIP: reportlab未インストール"
    WORKSPACE.mkdir(exist_ok=True)
    out = WORKSPACE / "test_report.pdf"
    c = rl_canvas.Canvas(str(out))
    c.drawString(100, 750, "Agent QA Report")
    c.drawString(100, 700, datetime.now().strftime("%Y-%m-%d %H:%M"))
    c.drawString(100, 650, "TEST PASSED")
    c.save()
    assert out.exists() and out.stat().st_size > 0
    return f"OK: {out.stat().st_size} bytes"


def test_P2():
    """PDFテキスト抽出"""
    try:
        import pdfplumber
    except ImportError:
        return "SKIP: pdfplumber未インストール"
    pdf = WORKSPACE / "test_report.pdf"
    if not pdf.exists():
        test_P1()
    if not pdf.exists():
        return "SKIP: PDFが生成されなかった"
    with pdfplumber.open(str(pdf)) as p:
        text = p.pages[0].extract_text() or ""
    assert len(text) > 0
    out = WORKSPACE / "extracted.txt"
    out.write_text(text)
    return f"OK: {len(text)} 文字抽出"


def test_P3():
    """PDF複数ページ生成"""
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
    except ImportError:
        return "SKIP: reportlab未インストール"
    WORKSPACE.mkdir(exist_ok=True)
    out = WORKSPACE / "multipage.pdf"
    c = rl_canvas.Canvas(str(out), pagesize=A4)
    for i in range(1, 4):
        c.drawString(100, 750, f"Page {i} of 3")
        c.showPage()
    c.save()
    assert out.exists() and out.stat().st_size > 0
    return f"OK: 3ページ PDF 生成 ({out.stat().st_size} bytes)"


def test_P4():
    """PDF結合"""
    try:
        from pypdf import PdfWriter, PdfReader
        from reportlab.pdfgen import canvas as rl_canvas
        import io
    except ImportError:
        return "SKIP: pypdf/reportlab未インストール"

    def _make_pdf_bytes(text: str) -> io.BytesIO:
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf)
        c.drawString(100, 750, text)
        c.save()
        buf.seek(0)
        return buf

    writer = PdfWriter()
    for text in ["Page A", "Page B"]:
        reader = PdfReader(_make_pdf_bytes(text))
        writer.add_page(reader.pages[0])
    out = WORKSPACE / "merged.pdf"
    with open(out, "wb") as f:
        writer.write(f)
    n_pages = len(writer.pages)
    assert n_pages == 2
    return f"OK: {n_pages} ページ結合"


PDF_TESTS = [
    {"id": "P1", "label": "PDF生成",         "fn": test_P1},
    {"id": "P2", "label": "PDFテキスト抽出",  "fn": test_P2},
    {"id": "P3", "label": "PDF複数ページ",    "fn": test_P3},
    {"id": "P4", "label": "PDF結合",          "fn": test_P4},
]


# =====================================================
# カテゴリ4: Excel（直接Python実行）
# =====================================================

def test_E1():
    """Excel生成"""
    try:
        import openpyxl
        from random import randint
    except ImportError:
        return "SKIP: openpyxl未インストール"
    WORKSPACE.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["月", "売上", "前月比"])
    prev = None
    for i in range(1, 7):
        sales = randint(10000, 50000)
        ratio = f"{sales / prev * 100:.1f}%" if prev else "-"
        ws.append([f"{i}月", sales, ratio])
        prev = sales
    out = WORKSPACE / "sales.xlsx"
    wb.save(str(out))
    assert out.exists()
    return f"OK: {out.stat().st_size} bytes"


def test_E2():
    """Excel集計"""
    try:
        import pandas as pd
    except ImportError:
        return "SKIP: pandas未インストール"
    xlsx = WORKSPACE / "sales.xlsx"
    if not xlsx.exists():
        test_E1()
    if not xlsx.exists():
        return "SKIP: sales.xlsxが生成されなかった"
    df = pd.read_excel(str(xlsx))
    col = df.columns[1]
    total = df[col].sum()
    avg = df[col].mean()
    assert total > 0
    return f"OK: 合計={total:.0f}, 平均={avg:.0f}"


def test_E3():
    """Excel→CSV変換"""
    try:
        import pandas as pd
    except ImportError:
        return "SKIP: pandas未インストール"
    xlsx = WORKSPACE / "sales.xlsx"
    if not xlsx.exists():
        test_E1()
    if not xlsx.exists():
        return "SKIP: sales.xlsxが生成されなかった"
    df = pd.read_excel(str(xlsx))
    out = WORKSPACE / "sales.csv"
    df.to_csv(str(out), index=False, encoding="utf-8-sig")
    assert out.exists()
    return f"OK: {len(df)} 行をCSV変換"


def test_E4():
    """Excelグラフ追加"""
    try:
        import openpyxl
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        return "SKIP: openpyxl未インストール"
    xlsx = WORKSPACE / "sales.xlsx"
    if not xlsx.exists():
        test_E1()
    if not xlsx.exists():
        return "SKIP: sales.xlsxが生成されなかった"
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb.active
    chart = BarChart()
    chart.title = "月別売上"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "E2")
    out = WORKSPACE / "sales_chart.xlsx"
    wb.save(str(out))
    assert out.exists()
    return f"OK: グラフ付きExcel {out.stat().st_size} bytes"


EXCEL_TESTS = [
    {"id": "E1", "label": "Excel生成",   "fn": test_E1},
    {"id": "E2", "label": "Excel集計",   "fn": test_E2},
    {"id": "E3", "label": "Excel→CSV",  "fn": test_E3},
    {"id": "E4", "label": "Excelグラフ", "fn": test_E4},
]


# =====================================================
# カテゴリ5: Web（直接Python実行）
# =====================================================

def test_W1():
    """HTTPリクエスト"""
    import requests
    r = requests.get("https://httpbin.org/json", timeout=15)
    assert r.status_code == 200
    data = r.json()
    return f"OK: keys={list(data.keys())}"


def test_W2():
    """HTML解析"""
    import requests
    from bs4 import BeautifulSoup
    # macOS の SSL 証明書バンドル問題を回避
    r = requests.get("https://example.com", timeout=15, verify=False)
    soup = BeautifulSoup(r.text, "html.parser")
    h1 = soup.find("h1")
    assert h1 is not None
    return f"OK: h1='{h1.text[:30]}'"


def test_W3():
    """JSON API取得・保存"""
    import requests
    r = requests.get("https://jsonplaceholder.typicode.com/todos/1", timeout=15)
    data = r.json()
    WORKSPACE.mkdir(exist_ok=True)
    out = WORKSPACE / "todo.json"
    out.write_text(json.dumps(data, indent=2))
    assert "title" in data
    return f"OK: title='{data['title'][:30]}'"


WEB_TESTS = [
    {"id": "W1", "label": "HTTPリクエスト", "fn": test_W1},
    {"id": "W2", "label": "HTML解析",       "fn": test_W2},
    {"id": "W3", "label": "JSON API保存",   "fn": test_W3},
]


# =====================================================
# テスト実行エンジン
# =====================================================

def run_direct_test(test: dict) -> dict:
    """エージェント不使用の直接テスト"""
    start = time.time()
    try:
        output = test["fn"]()
        skipped = str(output).startswith("SKIP")
        success = not skipped and "OK" in str(output)
        return {
            "id": test["id"],
            "label": test["label"],
            "success": success,
            "output": str(output),
            "elapsed": round(time.time() - start, 1),
            "retried": False,
            "skipped": skipped,
        }
    except Exception as e:
        import traceback
        return {
            "id": test["id"],
            "label": test["label"],
            "success": False,
            "output": f"ERROR: {e}\n{traceback.format_exc()[-300:]}",
            "elapsed": round(time.time() - start, 1),
            "retried": False,
            "skipped": False,
        }


def run_agent_test(test: dict) -> dict:
    """エージェント経由テスト（サブプロセス、タイムアウト付き）"""
    import shutil
    # ワークスペースをクリア
    if WORKSPACE.exists():
        shutil.rmtree(WORKSPACE)
    WORKSPACE.mkdir(parents=True)

    max_steps = test.get("max_steps", 5)
    timeout = TIMEOUT

    start = time.time()
    try:
        proc = subprocess.run(
            [VENV_PYTHON, MAIN_PY],
            input=test["task"],
            capture_output=True,
            text=True,
            cwd=str(AGENT_ROOT),
            timeout=timeout,
            env={**os.environ, "PYTHONPATH": str(AGENT_ROOT)},
        )
        out = proc.stdout + proc.stderr
        timed_out = False
    except subprocess.TimeoutExpired as e:
        raw = (e.stdout or b"")
        out = raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else (raw or "")
        out += f"\n[TIMEOUT after {timeout}s]"
        timed_out = True

    elapsed = round(time.time() - start, 1)
    expect = test.get("expect", "")

    # 成功判定: 結果行か exit code 0 ブロックに期待値を探す
    import re as _re
    success = False
    lines = out.splitlines()
    for ln in lines:
        if ln.startswith("結果:") and expect in ln:
            success = True
            break
    if not success:
        for i, ln in enumerate(lines):
            if "[exit code 0]" in ln:
                block = "\n".join(lines[i: i + 20])
                if expect in block:
                    success = True
                    break
    if not success and _re.search(r"^完了$", out, _re.MULTILINE) and expect in out:
        success = True

    return {
        "id": test["id"],
        "label": test["label"],
        "success": success,
        "output": out[-400:],
        "elapsed": elapsed,
        "retried": False,
        "skipped": False,
        "timed_out": timed_out,
    }


def _warmup_ollama() -> bool:
    """Codingテスト前に Ollama のモデルをメモリにロードしておく。"""
    import urllib.request, json as _json
    try:
        sys.path.insert(0, str(AGENT_ROOT))
        import llm as _llm
        model = getattr(_llm, "CODER_MODEL", "qwen2.5-coder:7b")
    except Exception:
        model = "qwen2.5-coder:7b"

    print(f"  🔥 Ollama ウォームアップ中 (model={model}) ...", end=" ", flush=True)
    payload = json.dumps({
        "model": model,
        "prompt": "hi",
        "stream": False,
        "options": {"num_predict": 1},
    }).encode()
    try:
        req = urllib.request.Request(
            "http://localhost:11434/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=300) as resp:
            _ = resp.read()
        print("✅ ロード完了")
        return True
    except Exception as e:
        print(f"⚠️  失敗: {e}")
        return False


def run_all(retry_failed: bool = True) -> list[dict]:
    WORKSPACE.mkdir(exist_ok=True)
    all_results: list[dict] = []

    categories = [
        ("Coding", CODING_TESTS, True),
        ("File",   FILE_TESTS,   False),
        ("PDF",    PDF_TESTS,    False),
        ("Excel",  EXCEL_TESTS,  False),
        ("Web",    WEB_TESTS,    False),
    ]

    for cat_name, tests, use_agent in categories:
        print(f"\n{'─' * 50}")
        print(f"  {cat_name}")
        print(f"{'─' * 50}")
        if use_agent:
            _warmup_ollama()

        for t in tests:
            print(f"  [{t['id']}] {t['label']} ...", end=" ", flush=True)

            r = run_agent_test(t) if use_agent else run_direct_test(t)
            r["category"] = cat_name

            if not r["success"] and not r.get("skipped") and retry_failed:
                print(f"❌ → リトライ中...", end=" ", flush=True)
                r2 = run_agent_test(t) if use_agent else run_direct_test(t)
                r2["category"] = cat_name
                if r2["success"]:
                    r2["retried"] = True
                    r = r2

            status = "✅" if r["success"] else ("⏭️ SKIP" if r.get("skipped") else "❌")
            print(f"{status} ({r['elapsed']}s)")
            if not r["success"] and not r.get("skipped"):
                snippet = r["output"].splitlines()
                for ln in snippet[-3:]:
                    print(f"     └─ {ln[:120]}")

            all_results.append(r)

    return all_results


# =====================================================
# レポート生成
# =====================================================

def save_report(results: list[dict]) -> Path:
    passed  = sum(1 for r in results if r["success"])
    skipped = sum(1 for r in results if r.get("skipped"))
    failed  = len(results) - passed - skipped
    retried = sum(1 for r in results if r.get("retried"))
    total_t = round(sum(r["elapsed"] for r in results), 1)

    try:
        sys.path.insert(0, str(AGENT_ROOT))
        import llm
        model = getattr(llm, "CODER_MODEL", "unknown")
    except Exception:
        model = "unknown"

    lines = [
        "# QA Report (Fast)",
        "",
        f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Model: {model}",
        "",
        "## Summary",
        "",
        "| 指標 | 値 |",
        "|------|-----|",
        f"| 総テスト数 | {len(results)} |",
        f"| 成功 | {passed} ({passed / len(results) * 100:.0f}%) |",
        f"| 失敗 | {failed} |",
        f"| スキップ | {skipped} |",
        f"| リトライで成功 | {retried} |",
        f"| 総実行時間 | {total_t}s |",
        "",
    ]

    for cat in ["Coding", "File", "PDF", "Excel", "Web"]:
        cat_r = [r for r in results if r["category"] == cat]
        if not cat_r:
            continue
        ok = sum(1 for r in cat_r if r["success"])
        lines += [
            f"## {cat} ({ok}/{len(cat_r)})",
            "",
            "| ID | Label | 結果 | 時間 | リトライ |",
            "|----|-------|------|------|---------|",
        ]
        for r in cat_r:
            st = "✅" if r["success"] else ("⏭️ SKIP" if r.get("skipped") else "❌")
            retry = "✅" if r.get("retried") else "-"
            lines.append(f"| {r['id']} | {r['label']} | {st} | {r['elapsed']}s | {retry} |")
        lines.append("")

    fails = [r for r in results if not r["success"] and not r.get("skipped")]
    if fails:
        lines += ["## 失敗テスト (要対応)", ""]
        for r in fails:
            lines += [
                f"### {r['id']}: {r['label']}",
                "",
                "```",
                r["output"][-400:],
                "```",
                "",
            ]
    else:
        lines += ["## 失敗テスト", "", "すべて成功しました。", ""]

    report_path = AGENT_ROOT / "qa_report_fast.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


# =====================================================
# ENTRYPOINT
# =====================================================

if __name__ == "__main__":
    print("🚀 Fast QA Loop 開始")
    print(f"   タイムアウト: {TIMEOUT}秒/テスト")
    print(f"   エージェント使用: Codingのみ (max_steps=5)")
    print(f"   直接実行: File / PDF / Excel / Web\n")

    results = run_all(retry_failed=True)

    report_path = save_report(results)
    passed = sum(1 for r in results if r["success"])

    print(f"\n{'=' * 50}")
    print(f"  最終結果: {passed}/{len(results)} 成功")
    print(f"{'=' * 50}")
    print(f"\n📄 レポート保存: {report_path}")
    print()
    print(report_path.read_text(encoding="utf-8"))
