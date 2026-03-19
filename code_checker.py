"""
code_checker.py — スキル獲得・toolkit登録時の自動コード品質チェック。

チェック項目:
  1. 構文チェック (ast.parse)
  2. Excel公式をPythonコードに混入していないか (RULE #8)
  3. .xlsx/.pdf 等のバイナリを python で直接実行しようとしていないか
  4. os.system/subprocess + f-string によるシェルインジェクションリスク
  5. bare except: (Exception を指定しない)
  6. toolkit 関数名が tool_ プレフィックスを持つか
  7. 空関数 (pass / ... のみ)

使い方:
  from code_checker import check_code, check_history_code, format_report

  issues = check_code(code_str, source_label="make_excel.py")
  issues = check_history_code(history, task_description)
  print(format_report(issues, source="make_excel.py"))

戻り値: list[dict]  例: [{"level": "error"|"warn", "rule": str, "msg": str}]
"""

from __future__ import annotations

import ast
import re
from typing import Any


# =====================================================
# 公開 API
# =====================================================

def check_code(code: str, source_label: str = "") -> list[dict]:
    """
    Python コード文字列を受け取り、問題リストを返す。
    """
    problems: list[dict] = []

    _check_syntax(code, source_label, problems)
    _check_excel_formulas(code, source_label, problems)
    _check_run_binary(code, source_label, problems)
    _check_shell_injection(code, source_label, problems)
    _check_bare_except(code, source_label, problems)
    _check_async_in_sync(code, source_label, problems)
    _check_duplicate_except(code, source_label, problems)
    _check_empty_function(code, source_label, problems)
    _check_toolkit_prefix(code, source_label, problems)

    return problems


def check_history_code(
    history: list[dict],
    task_description: str = "",
) -> list[dict]:
    """
    main.py の history リストを走査し、create_file / edit_file の
    .py コンテンツ全てを check_code でチェックする。
    """
    all_problems: list[dict] = []
    seen: set[str] = set()

    for h in history:
        action = h.get("action", {})
        if action.get("tool") not in ("create_file", "edit_file", "append_file"):
            continue
        path = action.get("path", "")
        if not path.endswith(".py"):
            continue
        content = action.get("content", "") or ""
        if not content or content in seen:
            continue
        seen.add(content)
        label = path or task_description[:40]
        problems = check_code(content, source_label=label)
        all_problems.extend(problems)

    return all_problems


def format_report(problems: list[dict], source: str = "") -> str:
    """問題リストを人間が読みやすい文字列に変換する。"""
    if not problems:
        return ""
    prefix = f"[{source}] " if source else ""
    lines = [f"    🔍 コードチェック: {prefix}{len(problems)}件の問題"]
    for p in problems:
        icon = "❌" if p["level"] == "error" else "⚠️"
        lines.append(f"      {icon} [{p['rule']}] {p['msg']}")
    return "\n".join(lines)


# =====================================================
# 個別チェック
# =====================================================

def _check_syntax(code: str, label: str, out: list[dict]) -> None:
    """ast.parse で構文エラーを検出する。"""
    try:
        ast.parse(code)
    except SyntaxError as e:
        out.append({
            "level": "error",
            "rule": "SYNTAX",
            "msg": f"{label}: 構文エラー — {e.msg} (行 {e.lineno})",
        })


# Excel関数名パターン (大文字)
_EXCEL_FUNCS = (
    "SUM|AVERAGE|IF|IFERROR|VLOOKUP|HLOOKUP|INDEX|MATCH|COUNT|COUNTA|COUNTIF|"
    "MAX|MIN|ROUND|ABS|INT|MOD|AND|OR|NOT|LEN|LEFT|RIGHT|MID|TRIM|"
    "CONCATENATE|CONCAT|TODAY|NOW|DATE|YEAR|MONTH|DAY|OFFSET|INDIRECT"
)

# 文字列リテラルの開始が = のあと Excel 関数名 or セル参照
_EXCEL_IN_STR = re.compile(
    r'["\']='                              # 文字列開始直後に =
    r'(?:'
    r'(?:' + _EXCEL_FUNCS + r')\s*\('     # Excel関数名( 例: =SUM(
    r'|[A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?'  # セル参照 例: =A1 or =B2:B7
    r')',
    re.IGNORECASE,
)


def _check_excel_formulas(code: str, label: str, out: list[dict]) -> None:
    """
    Python コード内に Excel 公式文字列が混入していないか検出する。
    例: ws["B1"] = "=SUM(B2:B7)" や data = ["=A1*100"]
    """
    m = _EXCEL_IN_STR.search(code)
    if m:
        # コメント行か確認 (# から始まる行はスキップ)
        start = m.start()
        line_start = code.rfind("\n", 0, start) + 1
        line_text = code[line_start:code.find("\n", start)]
        if not line_text.lstrip().startswith("#"):
            snippet = m.group(0)[:30]
            out.append({
                "level": "error",
                "rule": "RULE-8",
                "msg": f"{label}: Pythonコード内にExcel公式を検出 → {snippet!r} (文字列をExcel公式にしない)",
            })


_BINARY_PYTHON_RUN = re.compile(
    r'python[23]?\s+[\w./ \\-]*\.(xlsx|xls|pdf|docx|db|png|jpg|jpeg|gif|zip)',
    re.IGNORECASE,
)


def _check_run_binary(code: str, label: str, out: list[dict]) -> None:
    """
    python *.xlsx / *.pdf 等のバイナリファイルを実行しようとしているか検出。
    例: os.system("python sales.xlsx"), subprocess.run(["python", "report.pdf"])
    """
    m = _BINARY_PYTHON_RUN.search(code)
    if m:
        out.append({
            "level": "error",
            "rule": "RULE-8",
            "msg": f"{label}: バイナリファイルをpythonで実行しようとしている → {m.group(0)!r}",
        })


def _check_shell_injection(code: str, label: str, out: list[dict]) -> None:
    """
    shell=True / os.system + f-string や .format() の組み合わせを検出。
    """
    # shell=True を含む行の前後3行を確認
    lines = code.splitlines()
    risky_markers = ("shell=True", "os.system(", "os.popen(")

    for i, line in enumerate(lines):
        if not any(m in line for m in risky_markers):
            continue
        context = "\n".join(lines[max(0, i - 2):i + 3])
        if re.search(r'f["\'].*\{|\.format\s*\(|%\s*\(', context):
            out.append({
                "level": "warn",
                "rule": "SHELL-INJ",
                "msg": (
                    f"{label}: シェル実行 + 文字列フォーマットの組み合わせ "
                    f"(インジェクションリスク) — 行{i+1}"
                ),
            })
            break


def _check_bare_except(code: str, label: str, out: list[dict]) -> None:
    """bare `except:` (例外型を指定しない) を検出する。"""
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return

    for node in ast.walk(tree):
        if isinstance(node, ast.ExceptHandler) and node.type is None:
            out.append({
                "level": "warn",
                "rule": "BARE-EXCEPT",
                "msg": f"{label}: bare `except:` を検出 — `except Exception:` への変更を推奨",
            })
            break


def _check_empty_function(code: str, label: str, out: list[dict]) -> None:
    """関数本体が pass / ... / docstring のみの空関数を検出する。"""
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return

    for node in ast.walk(tree):
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        non_trivial = [
            n for n in node.body
            if not (
                isinstance(n, ast.Pass)
                or (isinstance(n, ast.Expr) and isinstance(n.value, ast.Constant)
                    and n.value.value is ...)
                or (isinstance(n, ast.Expr) and isinstance(n.value, ast.Constant)
                    and isinstance(n.value.value, str))  # docstring
            )
        ]
        if not non_trivial:
            out.append({
                "level": "warn",
                "rule": "EMPTY-FUNC",
                "msg": f"{label}: 空関数 `{node.name}()` — 本体なし (pass/... のみ)",
            })


def _check_async_in_sync(code: str, label: str, out: list[dict]) -> None:
    """
    通常の def (非async) の直接ボディで await / async with / async for を使っていないか検出する。
    ネストされた async def の中は除外する。
    """
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return

    def _walk_skip_async_def(node):
        """ast.walk と同様だが、ネストされた AsyncFunctionDef には降りない。"""
        for child in ast.iter_child_nodes(node):
            if isinstance(child, ast.AsyncFunctionDef):
                continue  # ネストされた async def の内部はスキップ
            yield child
            yield from _walk_skip_async_def(child)

    def _direct_async_usage(func_node: ast.FunctionDef):
        """func_node 直下のボディで async 構文を使っている箇所を返す。"""
        for child in _walk_skip_async_def(func_node):
            if isinstance(child, (ast.AsyncWith, ast.AsyncFor, ast.Await)):
                return child
        return None

    for node in ast.walk(tree):
        if not isinstance(node, ast.FunctionDef):
            continue
        # ネストされた async def を除いた直接ボディを確認
        problem = _direct_async_usage(node)
        if problem:
            kind = {
                ast.AsyncWith: "async with",
                ast.AsyncFor:  "async for",
                ast.Await:     "await",
            }.get(type(problem), "async 構文")
            line = getattr(problem, "lineno", "?")
            out.append({
                "level": "error",
                "rule": "ASYNC-IN-SYNC",
                "msg": (
                    f"{label}: 非async def `{node.name}()` 直下で "
                    f"`{kind}` を使用 → 呼び出し時にSyntaxError (行{line})"
                ),
            })
            break


def _check_duplicate_except(code: str, label: str, out: list[dict]) -> None:
    """
    同一 try ブロック内に重複する except Exception ハンドラや
    到達不能な except ハンドラを検出する。
    """
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return

    for node in ast.walk(tree):
        if not isinstance(node, ast.Try):
            continue
        handlers = node.handlers
        seen_types: list[str] = []
        for h in handlers:
            if h.type is None:
                type_str = "bare"
            else:
                type_str = ast.unparse(h.type) if hasattr(ast, "unparse") else getattr(h.type, "id", "?")
            # Exception より広い型が既に出た後の handler は到達不能
            if "Exception" in seen_types or "BaseException" in seen_types:
                out.append({
                    "level": "warn",
                    "rule": "DEAD-EXCEPT",
                    "msg": (
                        f"{label}: `except {type_str}` は到達不能 "
                        f"— 前の `except Exception` がすべてを捕捉する (行{h.lineno})"
                    ),
                })
                break
            # 同一型の重複
            if type_str in seen_types:
                out.append({
                    "level": "warn",
                    "rule": "DUP-EXCEPT",
                    "msg": f"{label}: `except {type_str}` が重複 (行{h.lineno})",
                })
                break
            seen_types.append(type_str)


def _check_toolkit_prefix(code: str, label: str, out: list[dict]) -> None:
    """
    toolkit / evolved ディレクトリのコードで、
    関数名が tool_ プレフィックスを持つか確認する。
    """
    if "toolkit" not in label.lower() and "evolved" not in label.lower():
        return
    funcs = re.findall(r"def (\w+)\(", code)
    public_funcs = [f for f in funcs if not f.startswith("_")]
    if public_funcs and not any(f.startswith("tool_") for f in public_funcs):
        out.append({
            "level": "warn",
            "rule": "TOOLKIT-PREFIX",
            "msg": f"{label}: toolkit関数名に `tool_` プレフィックスがない — {public_funcs[:3]}",
        })


# =====================================================
# 簡易セルフテスト
# =====================================================

if __name__ == "__main__":
    tests = [
        (
            "bad_excel.py",
            '''\
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "Month"
ws["B1"] = "=SUM(B2:B7)"
wb.save("sales.xlsx")
''',
            ["RULE-8"],  # expected rules
        ),
        (
            "bad_syntax.py",
            "def foo(\n    pass\n",
            ["SYNTAX"],
        ),
        (
            "bad_shell.py",
            '''\
import subprocess
cmd = f"grep {user_query} file.txt"
subprocess.run(cmd, shell=True)
''',
            ["SHELL-INJ"],
        ),
        (
            "bad_binary_run.py",
            '''\
import os
os.system("python sales.xlsx")
''',
            ["RULE-8"],
        ),
        (
            "bad_bare_except.py",
            '''\
try:
    x = 1/0
except:
    pass
''',
            ["BARE-EXCEPT"],
        ),
        (
            "good_excel.py",
            '''\
import openpyxl
import random

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Month", "Sales", "MoM"])
    for i in range(1, 7):
        ws.append([i, random.randint(10000, 50000), 0.05])
    wb.save("sales.xlsx")

if __name__ == "__main__":
    create_excel()
    print("Done")
''',
            [],  # no issues expected
        ),
    ]

    all_passed = True
    for label, code, expected_rules in tests:
        issues = check_code(code, label)
        found_rules = [i["rule"] for i in issues]
        ok = all(r in found_rules for r in expected_rules) and (
            len(issues) == 0 if not expected_rules else True
        )
        status = "✅" if ok else "❌"
        if not ok:
            all_passed = False
        print(f"{status} {label}: expected={expected_rules} got={found_rules}")
        if issues:
            print(format_report(issues, label))

    print()
    print("✅ 全テスト通過" if all_passed else "❌ 失敗あり")
